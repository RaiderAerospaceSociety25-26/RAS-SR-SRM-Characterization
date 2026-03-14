# ------------------ IMPORT ----------------
import os
import pandas as pd
from TestFire import TestFire
from Calculation import Calculation
from scipy.optimize import fsolve
from scipy.optimize import curve_fit
from scipy.ndimage import gaussian_filter1d
import numpy as np
import math
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook

# ----------------------------- GLOBAL VARIABLES ----------------------------
configFilename = '../dat/sample_test_config.txt' #INPUT: MUST SET THIS FILENAME TO THE NAME OF THE CONFIGURATION FILE
dataFilenameBase = '../dat/' #filename base for the Excel sheet; leave as a blank string except for any required prefix (default: '../dat/')    
testFires = [] #array to hold TestFire objects


# ---------------- READ DATA AND CREATE TESTFIRE OBJECTS ------------------------
def throwParseError():
    print("There is a configuration file formatting problem.")
    raise SystemExit(1)

def readImportOptions():
    #read import options from .txt file
    configFile = open(configFilename, 'r') #open config file to read
    configLines = [] #define array to store the config file data
    configLinesEdited = [] #define array (2D) to store the edited config file data
    i = 0
    for line in configFile:
        configLines.append(line.strip()) #add each file line to the configLines array
    configFile.close() #close file
    startIndices = []
    endIndices = []
    i = 0
    for line in configLines:
        if line == "!CONFIGSTART": startIndices.append(i) #index of the array where "!CONFIGSTART" is stored
        if line == "!CONFIGEND": endIndices.append(i) #index of the array where "!CONFIGEND" is stored
        i += 1
    if (len(startIndices) != len(endIndices)) or (len(startIndices) == 0) or (len(endIndices) == 0): throwParseError() #this means that the number of !CONFIGSTART or !CONFIGEND entries is wrong
    for j in range(len(startIndices)):
        configLinesEdited.append(configLines[startIndices[j]+1:endIndices[j]]) #stores the configuration data without the !CONFIGSTART and !CONFIGEND tags
    #results in configLinesEdited being an array of an array of strings with each string element being a line of options
    return configLinesEdited

#function to smooth data. inputs are data (array containing data you want to smooth), time (time array associated with the data array), and window_duration (number controlling how much smoothing to do)
def smoothData(data, time, window_duration):
    if window_duration == -1: return data

    dt = np.diff(time)
    avg_dt = np.mean(dt)
    sampling_rate = 1 / avg_dt  # Hz

    # Convert window duration (in seconds) to sigma in sample units
    samples_per_window = window_duration / avg_dt
    sigma_samples = samples_per_window / 2.355  # FWHM ≈ 2.355σ for Gaussian

    # Apply Gaussian filter
    smoothed_data = gaussian_filter1d(data, sigma=sigma_samples)

    return smoothed_data


#creates each TestFire object and sets the import options for each using the allConfigs array
def setImportOptions(allConfigs):
    for x in range(len(allConfigs)):
        fileEntry = allConfigs[x][0].split(" ")
        psmoothingEntry = allConfigs[x][1].split(" ")
        tsmoothingEntry = allConfigs[x][2].split(" ")
        geomEntry = allConfigs[x][3].split(" ")
        throatEntry = allConfigs[x][4].split(" ")
        massEntry = allConfigs[x][5].split(" ")
        densityEntry = allConfigs[x][6].split(" ")
        arr = []
        for y in range(len(throatEntry)-2): #for each fire, set filename, pressure units, thrust units, geometry units, geometry, throat units, and throat
            testFire = TestFire()
            testFire.set_fireIndex(y)
            if (fileEntry[0] == "filename"):
                testFire.set_filename(fileEntry[1])
                #set sheetnames
                excelFile = pd.ExcelFile(dataFilenameBase + testFire.get_filename()) #define pandas ExcelFile object
                sheetnames = excelFile.sheet_names #get sheet names in Excel workbook
                testFire.set_sheetnames(sheetnames)
                testFire.set_pressUnits(fileEntry[2])
                testFire.set_thrustUnits(fileEntry[3])
            else: throwParseError()
            if (psmoothingEntry[0] == "psmoothing"):
                if psmoothingEntry[1] == "off":
                    testFire.set_psmoothing(False)
                elif psmoothingEntry[1] == "on":
                    testFire.set_psmoothing(True)
                    numArr = []
                    for i in range(2, len(psmoothingEntry)):
                        numArr.append(float(psmoothingEntry[i]))
                    testFire.set_psmoothingArr(numArr)
                else: throwParseError()
            else: throwParseError()
            if (tsmoothingEntry[0] =="tsmoothing"):
                if tsmoothingEntry[1] == "off": testFire.set_tsmoothing(False)
                elif tsmoothingEntry[1] == "on":
                    testFire.set_tsmoothing(True)
                    numArr = []
                    for i in range(2, len(tsmoothingEntry)):
                        numArr.append(float(tsmoothingEntry[i]))
                    testFire.set_tsmoothingArr(numArr)
                else: throwParseError()
            else: throwParseError()
            if (geomEntry[0] == "geometry"):
                testFire.set_geomUnits(geomEntry[1])
                geometryArr = geomEntry[2].split("||")
                geometryArrFinal = []
                for list in geometryArr:
                    numArr = []
                    for num in list.split(","):
                        numArr.append(num)
                    geometryArrFinal.append(numArr)
                testFire.set_geometry(geometryArrFinal) #geometry array is of the format [[grain length, core diameter, grain OD], [grain length, core diameter, grain OD]] from forwardmost grain to aftmost grain
            else: throwParseError()
            if(throatEntry[0] == "throat"):
                testFire.set_throatUnits(throatEntry[1])
                testFire.set_throat(float(throatEntry[y+2]))
            else: throwParseError()
            if(massEntry[0] == "mass"):
                testFire.set_massUnits(massEntry[1])
                testFire.set_mass(float(massEntry[y+2]))
            else: throwParseError()
            if(densityEntry[0] == "density"):
                testFire.set_densityUnits(densityEntry[1])
                testFire.set_density(float(densityEntry[y+2]))
            else: throwParseError()
            arr.append(testFire)
        testFires.append(arr)

#reads data from Excel and adds it to each TestFire object
def readExcelData(allConfigs):
    x = 0
    for tfList in testFires:
        y = 0
        for tf in tfList:
            print("Reading data from input Excel to: " + str(testFires[x][y]))
            tfObj = testFires[x][y]
            testFires[x][y].set_dat(pd.read_excel((dataFilenameBase + tfObj.get_filename()), tfObj.get_sheetnames()[y]))
            y += 1
        x += 1


# --------------------- CHARACTERIZATION --------------------------
#take the data associated with each TestFire object and do necessary conversions and calculations to define a Calculation object, adding it to the TestFire object
#a Calculation object contains all the data necessary for characterization. in addition, all data is in metric units so there is no unit confusion during the characterization calculations.
def conversionsAndDefinitions():
    x = 0
    for tfList in testFires:
        y = 0
        for tf in tfList:
            data = testFires[x][y].get_dat()

            #burn_time
            burn_time = data.iloc[len(data.iloc[:,0])-1,0]

            #time
            time = data.iloc[:,0]

            #pressure
            if (tf.get_pressUnits() == "psig"): press = data.iloc[:,1]*6894.76 #convert psi to Pa
            elif (tf.get_pressUnits() == "kPa"): press = data.iloc[:,1]*1000 #convert kPa to Pa
            else: throwParseError()

            #smoothed pressure
            if (tf.get_psmoothing() == False): press_smoothed = None
            elif (tf.get_psmoothing() == True):
                press_smoothed = smoothData(press, time, tf.get_psmoothingArr()[y])

            #thrust
            if (tf.get_thrustUnits() == "N"): thrust = data.iloc[:,2]
            elif (tf.get_thrustUnits() == "lbf"): thrust = data.iloc[:,2]*4.44822 #convert lbf to N
            else: throwParseError()

            #smoothed thrust
            if (tf.get_tsmoothing() == False): thrust_smoothed = None
            elif (tf.get_tsmoothing() == True):
                thrust_smoothed = smoothData(thrust, time, tf.get_tsmoothingArr()[y])

            #mass
            if (tf.get_thrustUnits() == "N"): mass = tf.get_mass()
            elif (tf.get_massUnits() == "lb"): mass = tf.get_mass()*0.453592 #convert lbm to kg
            else: throwParseError()

            #density
            if (tf.get_densityUnits() == "kg/m3"): density = tf.get_density()
            elif (tf.get_densityUnits() == "lb/in3"): density = tf.get_density()*27679.9 #convert lbm/in^3 to kg/m^3
            else: throwParseError()
            
            #throat_area
            if (tf.get_throatUnits() == "mm"): throat_area = math.pi*(math.pow(tf.get_throat()*1000,2)/4) #convert to m^2
            elif (tf.get_throatUnits() == "in"): throat_area = math.pi*(math.pow((tf.get_throat()*0.0254),2)/4) #convert to m^2
            else: throwParseError()

            #grain_length
            grain_len = []
            for grain in tf.get_geometry():
                if (tf.get_geomUnits() == "mm"): grain_len.append(float(grain[0]))
                elif (tf.get_geomUnits() == "in"): grain_len.append(float(grain[0])*0.0254) #convert to m
                else: throwParseError()

            #grain_init_core
            grain_init_core = []
            for grain in tf.get_geometry():
                if (tf.get_geomUnits() == "mm"): grain_init_core.append(float(grain[1]))
                elif (tf.get_geomUnits() == "in"): grain_init_core.append(float(grain[1])*0.0254) #convert to m
                else: throwParseError()

            #grain_OD
            grain_OD = []
            for grain in tf.get_geometry():
                if (tf.get_geomUnits() == "m"): grain_OD.append(float(grain[2]))
                elif (tf.get_geomUnits() == "in"): grain_OD.append(float(grain[2])*0.0254) #convert to m
                else: throwParseError()

            #impulse
            if (tf.get_tsmoothing() == True): impulse = np.trapezoid(thrust_smoothed, time)
            else: impulse = np.trapezoid(thrust, time)

            #press_integral
            if (tf.get_psmoothing() == True): press_integral = np.trapezoid(press_smoothed, time)
            else: press_integral = np.trapezoid(press, time)

            #c_star
            c_star = (throat_area/mass)*press_integral

            #isp
            isp = impulse/(mass*9.81)

            #make Calculation object and add it to the proper TestFire object
            calc = Calculation(time, press, thrust, press_smoothed, thrust_smoothed, mass, density, throat_area, grain_len, grain_init_core, grain_OD, burn_time, impulse, press_integral, c_star, isp)
            testFires[x][y].set_calculation(calc)
            print("Set calculation " + str(calc) + " for " + str(testFires[x][y]))

            #increment
            y += 1
        x += 1


#function to perform all characterization calculations
#probably use either scipy.optimize.root_scalar() or scipy.optimize.fsolve()
def characterization():
    """
    function arguments:
    ds: instantaneous change in surface regression (variable that is changed to solve the equation)
    D: grain outer diameter
    d_0: grain initial core diameter
    L_0 = grain initial length
    rho_b = propellant density
    A_t = throat area
    dt = time step
    P = instantaneous pressure
    s_prev = the s value for the previous iteration of the function
    c_star = C*
    """
    def characterizationEqns(ds, D, d_0, L_0, rho_b, A_t, dt, P, s_prev, c_star):
        """
        additional variables:
        A_b = instantaneous burn area
        s = surface regression amount
        """
        s = s_prev + ds
        A_b = 0.0
        i = 0
        for num in D:
            A_b += math.pi*((0.5*((math.pow(D[i],2))-math.pow((d_0[i]+(2*s.item())),2)) + (L_0[i]-(2*s.item()))*(d_0[i]+(2*s.item()))))
            i += 1
        zero_func = ds - ((A_t*P*dt)/(A_b*rho_b*c_star))
        return zero_func

    x = 0
    for tfList in testFires:
        y = 0
        for tf in tfList:
            calc = testFires[x][y].get_calculation()
            dsArr = [] #array for ds
            sArr = [] #array for s
            A_bArr = [] #array for A_b
            ds_dt_arr = [] #array for ds/dt
            i = 0
            for line in calc.get_time():
                s_initial = 0.05 #initial guess for s
                D = calc.get_grain_OD()
                d_0 = calc.get_grain_init_core()
                L_0 = calc.get_grain_len()
                rho_b = calc.get_density()
                A_t = calc.get_throat_area()
                P = 0 # instantaneous pressure
                if tf.get_psmoothing() == True: P = calc.get_press_smoothed()[i]
                else: P = calc.get_press().iloc[i]
                c_star = calc.get_c_star()
                dt = 0
                s_prev = 0
                if (i != 0):
                    dt = calc.get_time().iloc[i] - calc.get_time().iloc[i-1]
                    s_prev = sArr[i-1]
                    ds_new = fsolve(characterizationEqns, s_initial, args=(D, d_0, L_0, rho_b, A_t, dt, P, s_prev, c_star))
                    dsArr.append(float(ds_new.item()))
                    sArr.append(sArr[i-1] + float(ds_new.item()))
                    ds_dt_arr.append(float(ds_new.item())/dt)
                else:
                    sArr.append(0)
                    dsArr.append(0)
                    ds_dt_arr.append(0)
                A_b = 0.0
                j = 0
                for num in D:
                    A_b += math.pi*((0.5*(math.pow(D[j],2))-math.pow((d_0[j]+(2*sArr[i])),2)) + ((L_0[j]-(2*sArr[i]))*(d_0[j]+(2*sArr[i])))) #SOMETHING IS THROWING AN ARRAY-RELATED ERROR HERE!!!
                    j += 1
                A_bArr.append(A_b)
                i += 1
            #add s, ds, A_b, and ds/dt to the calculation object
            testFires[x][y].get_calculation().set_A_bArr(A_bArr)
            testFires[x][y].get_calculation().set_sArr(sArr)
            testFires[x][y].get_calculation().set_dsArr(dsArr)
            testFires[x][y].get_calculation().set_ds_dtArr(ds_dt_arr)
            print("Finished characterization calculation for " + str(testFires[x][y]))
            y += 1
        print("Finished characterization calculation for propellant #" + str(x+1))
        x += 1

#function to perform conversions of important output data back to imperial and output to an Excel sheet
def outputResults():
    """
    figures to output (in both sets of units):
    thrust vs time and pressure vs time
    burn rate vs time and burn area vs time
    
    columns to output to excel sheet (original):
    time (original)
    thrust (original)
    pressure (original)

    more columns to output to excel sheet (from Calculation object, in both sets of units):
    time
    thrust
    pressure
    s
    ds
    A_b
    ds/dt

    additional info to output to excel sheet (in both sets of units):
    grain info: outer diameters, initial core diameters, grain initial lengths
    propellant density
    propellant mass
    throat diameter
    throat area
    pressure integral
    thrust integral
    C*
    ISP
    """
    #create folder for output files:
    #make "out1", "out2", etc., but back one folder, so you'll have three folders: "src", "dat", and "out1" after running
    outputFolderNum = 1
    outputFileBase = "../out/out"
    outputFileDir = ""
    while os.path.exists(outputFileBase + str(outputFolderNum)):
        outputFolderNum += 1
    outputFileDir = outputFileBase + str(outputFolderNum)
    os.makedirs(outputFileDir)

    x = 0
    for tfList in testFires:
        #create Excel file and ExcelWriter object
        wb = Workbook()
        file_path = outputFileDir + "/out"+str(x)+".xlsx"
        wb.save(file_path)
        excelWriter = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') #REQUIRES OPENPYXL LIBRARY!!!

        #lists for average pressure and average burn rate for characterization
        avgPress = []
        avgPressImperial = []
        avgBurnrate = []
        avgBurnrateImperial = []
        avgPressHigh = []
        avgPressImperialHigh = []
        avgBurnrateHigh = []
        avgBurnrateImperialHigh = []

        y = 0
        for tf in tfList:
            calc = testFires[x][y].get_calculation()
            outputFileBase2 = outputFileDir + "/" + str(x) + "," + str(y) + "_"

            time = calc.get_time() #sec
            thrust_original = calc.get_thrust() #N
            thrust = thrust_original
            if tf.get_tsmoothing() == True: thrust = calc.get_thrust_smoothed()
            press_original = calc.get_press()/1000000 #convert to MPa
            press = press_original
            if tf.get_psmoothing() == True: press = calc.get_press_smoothed()/1000000 #convert to MPa
            s = calc.get_sArr()
            ds = calc.get_dsArr()
            A_b = calc.get_A_bArr()
            ds_dt = [i * 1000 for i in calc.get_ds_dtArr()] #convert to mm/sec
            thrust_imperial = thrust*0.224809 #convert to lbf
            press_imperial = press*145.03774 #convert to psig
            s_imperial = [i * 39.3701 for i in calc.get_sArr()] #convert to inches
            ds_imperial = [i * 39.3701 for i in calc.get_dsArr()] #convert to inches
            A_b_imperial = [i * 1550.0031 for i in calc.get_A_bArr()] #convert to in^2
            ds_dt_imperial = [i * 39.3701 for i in calc.get_ds_dtArr()] #convert to in/sec
            #note that everything other than time, thrust, pressure is in an array. time, thrust, and pressure are in DataFrames.

            #also calculate average pressure and average thrust and add it to the arrays

            #thrust vs time and pressure vs time (metric)
            fig, ax1 = plt.subplots(figsize=(13, 8))
            #plot thrust
            ax1.plot(time, thrust, marker='o', linestyle='-', color="b", label="Thrust")
            ax1.set_xlabel("Time (sec)")
            ax1.set_ylabel("Thrust (N)")
            handles1, labels1 = ax1.get_legend_handles_labels()
            #plot pressure
            ax2 = ax1.twinx() #creates another axis that shares the x-axis
            ax2.plot(time, press, marker='^', linestyle='-', color="r", label="Pressure")
            ax2.set_ylim(top=max(press)*1.6) #sets max for the ylim to be a bit higher
            ax2.set_ylabel("Pressure (MPa)")
            handles2, labels2 = ax2.get_legend_handles_labels()
            #show
            #plt.legend(ax1, ax2)
            all_handles = handles1 + handles2
            all_labels = labels1 + labels2
            plt.legend(all_handles, all_labels)
            plt.savefig(outputFileBase2 + "_thrust-pressure_metric")
            plt.close()

            #thrust vs time and pressure vs time (imperial)
            fig, ax1 = plt.subplots(figsize=(13, 8))
            #plot thrust
            ax1.plot(time, thrust_imperial, marker='o', linestyle='-', color="b", label="Thrust")
            ax1.set_xlabel("Time (sec)")
            ax1.set_ylabel("Thrust (lbf)")
            handles1, labels1 = ax1.get_legend_handles_labels()
            #plot pressure
            ax2 = ax1.twinx() #creates another axis that shares the x-axis
            ax2.plot(time, press_imperial, marker='^', linestyle='-', color="r", label="Pressure")
            ax2.set_ylim(top=max(press_imperial)*1.6) #sets max for the ylim to be a bit higher
            ax2.set_ylabel("Pressure (psig)")
            #show
            all_handles = handles1 + handles2
            all_labels = labels1 + labels2
            plt.legend(all_handles, all_labels)
            plt.savefig(outputFileBase2 + "_thrust-pressure_imperial")
            plt.close()

            #burn rate vs time and burn area vs time (metric)
            fig, ax1 = plt.subplots(figsize=(13, 8))
            #plot burn rate
            ax1.plot(time, ds_dt, marker='o', linestyle='-', color="b", label="Burn rate ds/dt")
            ax1.set_xlabel("Time (sec)")
            ax1.set_ylabel("Burn rate (mm/sec)")
            handles1, labels1 = ax1.get_legend_handles_labels()
            #plot burn area
            ax2 = ax1.twinx() #creates another axis that shares the x-axis
            #A_b = [i * 1000000 for i in calc.get_A_bArr()] #use for conversion to mm^2
            ax2.plot(time, A_b, marker='^', linestyle='-', color="r", label="Burn area A_b")
            #ax2.set_ylim(top=max(A_b)*2)
            #ax2.set_ylim(bottom=0)
            ax2.set_ylabel("Burn area (m^2)")
            handles2, labels2 = ax2.get_legend_handles_labels()
            #show
            all_handles = handles1 + handles2
            all_labels = labels1 + labels2
            plt.legend(all_handles, all_labels)
            plt.savefig(outputFileBase2 + "_burnrate-burnarea_metric")
            plt.close()

            #burn rate vs time and burn area vs time (imperial)
            fig, ax1 = plt.subplots(figsize=(13, 8))
            #plot burn rate
            ax1.plot(time, ds_dt_imperial, marker='o', linestyle='-', color="b", label="Burn rate ds/dt")
            ax1.set_xlabel("Time (sec)")
            ax1.set_ylabel("Burn rate (in/sec)")
            handles1, labels1 = ax1.get_legend_handles_labels()
            #plot burn area
            ax2 = ax1.twinx() #creates another axis that shares the x-axis
            ax2.plot(time, A_b_imperial, marker='^', linestyle='-', color="r", label="Burn area A_b")
            #ax2.set_ylim(top=max(A_b_imperial)*2)
            #ax2.set_ylim(bottom=0)
            ax2.set_ylabel("Burn area (in^2)")
            handles2, labels2 = ax2.get_legend_handles_labels()
            #show
            all_handles = handles1 + handles2
            all_labels = labels1 + labels2
            plt.legend(all_handles, all_labels)
            plt.savefig(outputFileBase2 + "_burnrate-burnarea_imperial")
            plt.close()

            #output to excel sheet
            finalDataFrame = testFires[x][y].get_dat().copy() #start with original data
            #concatenate the rest of the data with the original data
            finalDataFrame = pd.concat(
                [finalDataFrame, pd.DataFrame({"Time (sec)": time}),
                 pd.DataFrame({"Pressure (MPa)": press_original}),
                 pd.DataFrame({"Thrust (N)": thrust_original}),
                 pd.DataFrame({"Burn Area A_b (m^2)": A_b}),
                 pd.DataFrame({"Instantaneous Regression ds (m)": ds}),
                 pd.DataFrame({"Cumulative Regression s (m)": s}),
                 pd.DataFrame({"Instantaneous Burn Rate ds/dt (mm/sec)": ds_dt}),
                 pd.DataFrame({"Time (sec)": time}),
                 pd.DataFrame({"Pressure (psi)": press_imperial}),
                 pd.DataFrame({"Thrust (lbf)": thrust_imperial}),
                 pd.DataFrame({"Burn Area A_b (in^2)": A_b_imperial}),
                 pd.DataFrame({"Instantaneous Regression ds (in)": ds_imperial}),
                 pd.DataFrame({"Cumulative Regression s (in)": s_imperial}),
                 pd.DataFrame({"Instantaneous Burn Rate ds/dt (in/sec)": ds_dt_imperial})], axis=1)

            #write more data for the fire to the spreadsheet
            burn_time = calc.get_burn_time()
            impulse = calc.get_impulse() #N*s
            press_integral = calc.get_press_integral()/1000000 #MPa*s
            c_star = calc.get_c_star() #m/s? I think?
            isp = calc.get_isp() #sec
            impulse_imperial = impulse*0.224809
            press_integral_imperial = press_integral*145.03773773
            c_star_imperial = c_star*3.280839895
            full_avg_press = press_integral / burn_time #average pressure calculated using the full pressure integral
            full_avg_press_imperial = press_integral_imperial / burn_time
            # ---------------------- average pressure and average burnrate calculation using only points with a pressure greater than the average pressure calculated above using the full pressure integral
            #(note this method may have issues if a fire has extremely noisy pressure data or if there is a massive drop in pressure and then a subsequent massive increase in pressure that should be considered, HOWEVER, I think it should generally be more accurate)
            press_trimmed = np.array([])
            time_trimmed = np.array([])
            burnrate_trimmed = np.array([])
            burnrate_trimmed_imperial = np.array([])
            started_adding = False
            i = 0
            for p in press:
                if p < full_avg_press and started_adding: break
                if p > full_avg_press:
                    press_trimmed = np.append(press_trimmed, press[i])
                    time_trimmed = np.append(time_trimmed, time[i])
                    burnrate_trimmed = np.append(burnrate_trimmed, ds_dt[i])
                    burnrate_trimmed_imperial = np.append(burnrate_trimmed_imperial, ds_dt_imperial[i])
                    started_adding = True
                i += 1
            high_avg_press = np.trapezoid(press_trimmed, time_trimmed) / (time_trimmed[len(time_trimmed)-1] - time_trimmed[0])
            high_avg_press_imperial = high_avg_press*145.03773773
            high_avg_burnrate = np.trapezoid(burnrate_trimmed, time_trimmed) / (time_trimmed[len(time_trimmed)-1] - time_trimmed[0])
            high_avg_burnrate_imperial = np.trapezoid(burnrate_trimmed_imperial, time_trimmed) / (time_trimmed[len(time_trimmed)-1] - time_trimmed[0])
            # ----------------------------------------------------------------------------------------------------------------------------------- end avg pressure and average burnrate calculation
            burnrate_integral = np.trapezoid(ds_dt, time)
            burnrate_integral_imperial = np.trapezoid(ds_dt_imperial, time)
            avg_burnrate = burnrate_integral / burn_time
            avg_burnrate_imperial = burnrate_integral_imperial / burn_time
            if tf.get_throatUnits() == "mm":
                throat_diameter = tf.get_throat()
                throat_area = math.pow(throat_diameter, 2) * math.pi / 4
                throat_diameter_imperial = throat_diameter/25.4
                throat_area_imperial = throat_area/(math.pow(25.4, 2))
            elif tf.get_throatUnits() == "in":
                throat_diameter_imperial = tf.get_throat()
                throat_area_imperial = math.pow(throat_diameter_imperial, 2) * math.pi / 4
                throat_diameter = throat_diameter_imperial*25.4
                throat_area = math.pow(throat_diameter, 2) * math.pi / 4
            extra_data_arr = np.array([throat_area, throat_diameter, throat_area_imperial, throat_diameter_imperial, burn_time, impulse, press_integral, c_star, full_avg_press, avg_burnrate, high_avg_press, high_avg_burnrate, impulse_imperial, press_integral_imperial, c_star_imperial, full_avg_press_imperial, avg_burnrate_imperial, high_avg_press_imperial, high_avg_burnrate_imperial, isp])
            extra_data_arr_strings = np.array(["Throat Area (mm^2)", "Throat Diameter (mm)", "Throat Area (in^2)", "Throat Diameter (in)", "Burn Time (sec)", "Impulse (N*s)", "Full Pressure Integral (MPa*s)", "C* (m/s)", "Average Pressure (using full fire) (MPa)", "Average Burnrate (using full fire) (mm/s)", "Average Pressure (using points above average) (MPa)", "Average Burnrate (using points above average) (mm/s)", "Impulse (lbf*s)", "Full Pressure Integral (psig*s)", "C* (ft/s)", "Average Pressure (using full fire) (psig)", "Average Burnrate (in/s) (using full fire)", "Average Pressure (using points above average) (psig)", "Average Burnrate (using points above average) (in/s)", "ISP (sec)"])
            finalDataFrame = pd.concat([finalDataFrame, pd.DataFrame({"str": extra_data_arr_strings}), pd.DataFrame({"dat": extra_data_arr})], axis=1)

            #write final dataframe to excel
            finalDataFrame.to_excel(excelWriter, sheet_name=str(y))

            #add averages to the correct arrays for later characterization calculations
            avgPress.append(full_avg_press)
            avgPressImperial.append(full_avg_press_imperial)
            avgBurnrate.append(avg_burnrate)
            avgBurnrateImperial.append(avg_burnrate_imperial)
            avgPressHigh.append(high_avg_press)
            avgPressImperialHigh.append(high_avg_press_imperial)
            avgBurnrateHigh.append(high_avg_burnrate)
            avgBurnrateImperialHigh.append(high_avg_burnrate_imperial)

            y += 1
        
        #function for calculating burnrate
        def func_powerlaw(x, a, n):
            return a * x**n
        
        #power law fit for original average pressure and burnrate calculation
        params, covariance = curve_fit(func_powerlaw, avgPress, avgBurnrate, maxfev=6000)
        a, n = params
        residuals = avgBurnrate - func_powerlaw(np.array(avgPress), *params)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((avgBurnrate - np.mean(avgBurnrate))**2)
        r_squared = 1 - (ss_res / ss_tot)
        params_imperial, covariance_imperial = curve_fit(func_powerlaw, avgPressImperial, avgBurnrateImperial, maxfev=6000)
        a_imperial, n_imperial = params_imperial
        strings = np.array(["Metric (MPa and mm):", "a", "n", "Imperial (psig and in):", "a", "n", " ", "R^2"])
        data = np.array([None, a, n, None, a_imperial, n_imperial, None, r_squared])
        characterizationDataframe = pd.concat([pd.DataFrame({"str": strings}), pd.DataFrame({"dat": data})], axis=1)
        characterizationDataframe.to_excel(excelWriter, sheet_name="CharacterizationOld")

        #output characterization plots (for original average pressure and burnrate calculation)
        x_metric = np.linspace(0, 6, 300)
        y_metric = a * (x_metric**n)
        plt.figure(figsize=(13, 8))
        plt.plot(x_metric, y_metric, linestyle='--', label="Power Law Fit")
        plt.scatter(avgPress, avgBurnrate, marker='o', c='b', label="Average Burnrate vs. Average Pressure")
        plt.xlabel("Average Pressure (MPa)")
        plt.ylabel("Average Burn Rate (mm/sec)")
        plt.savefig(outputFileDir + "/characterization_metric_" + str(x))
        plt.close()

        x_imperial = np.linspace(0, 700, 300)
        y_imperial = a_imperial * (x_imperial**n_imperial)
        plt.figure(figsize=(13, 8))
        plt.plot(x_imperial, y_imperial, linestyle='--', label="Power Law Fit")
        plt.scatter(avgPressImperial, avgBurnrateImperial, marker='o', c='b', label="Average Burnrate vs. Average Pressure")
        plt.xlabel("Average Pressure (psig)")
        plt.ylabel("Average Burn Rate (in/sec)")
        plt.savefig(outputFileDir + "/characterization_imperial_" + str(x))
        plt.close()

        #power law fit for new average pressure and burnrate calculation
        params2, covariance2 = curve_fit(func_powerlaw, avgPressHigh, avgBurnrateHigh, maxfev=6000)
        a2, n2 = params2
        residuals = avgBurnrateHigh - func_powerlaw(np.array(avgPressHigh), *params2)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((avgBurnrateHigh - np.mean(avgBurnrateHigh))**2)
        r_squared2 = 1 - (ss_res / ss_tot)
        params2_imperial, covariance_imperial2 = curve_fit(func_powerlaw, avgPressImperialHigh, avgBurnrateImperialHigh, maxfev=6000)
        a2_imperial, n2_imperial = params2_imperial
        strings2 = np.array(["Metric (MPa and mm):", "a", "n", "Imperial (psig and in):", "a", "n", " ", "R^2"])
        data2 = np.array([None, a2, n2, None, a2_imperial, n2_imperial, None, r_squared2])
        characterizationDataframe = pd.concat([pd.DataFrame({"str": strings2}), pd.DataFrame({"dat": data2})], axis=1)
        characterizationDataframe.to_excel(excelWriter, sheet_name="CharacterizationNew")

        #output characterization plots (for updated average pressure and burnrate calculation)
        x_metric = np.linspace(0, 8, 300)
        y_metric = a2 * (x_metric**n2)
        plt.figure(figsize=(13, 8))
        plt.plot(x_metric, y_metric, linestyle='--', label="Power Law Fit")
        plt.scatter(avgPressHigh, avgBurnrateHigh, marker='o', c='b', label="Average Burnrate vs. Average Pressure")
        plt.xlabel("Average Pressure (MPa)")
        plt.ylabel("Average Burn Rate (mm/sec)")
        plt.savefig(outputFileDir + "/characterization_metric_new_" + str(x))
        plt.close()

        x_imperial = np.linspace(0, 1100, 300)
        y_imperial = a2_imperial * (x_imperial**n2_imperial)
        plt.figure(figsize=(13, 8))
        plt.plot(x_imperial, y_imperial, linestyle='--', label="Power Law Fit")
        plt.scatter(avgPressImperialHigh, avgBurnrateImperialHigh, marker='o', c='b', label="Average Burnrate vs. Average Pressure")
        plt.xlabel("Average Pressure (psig)")
        plt.ylabel("Average Burn Rate (in/sec)")
        plt.savefig(outputFileDir + "/characterization_imperial_new_" + str(x))
        plt.close()

        #TODO: delete sheet named "Sheet" in the excel workbook
        excelWriter.close()

        print("Wrote all data to Excel and output figures for propellant #" + str(x+1))

        x += 1


# --------------------- MAIN ---------------------
#debug helper function to print all attributes associated with each TestFire object
def printAllTestFireListAttributes():
    x = 0
    for tfList in testFires:
        y = 0
        for tf in tfList:
            TestFire.printAllAttributes(tf)
            y += 1
        x += 1

def main():
    #import data
    allConfigs = readImportOptions()
    setImportOptions(allConfigs)
    readExcelData(allConfigs)
    #printAllTestFireListAttributes()

    #characterization calculations
    conversionsAndDefinitions()
    characterization()
    
    #results
    outputResults()

main()